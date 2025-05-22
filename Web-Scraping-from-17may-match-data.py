

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

# Cricbuzz scorecard URL
url = "https://www.cricbuzz.com/live-cricket-scorecard/118541/uae-vs-ban-1st-t20i-bangladesh-tour-of-uae-2025"
headers = {"User-Agent": "Mozilla/5.0"}
response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.content, "html.parser")

# Find all innings blocks
innings_blocks = soup.find_all("div", class_="cb-col cb-col-100 cb-ltst-wgt-hdr")

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

for i, block in enumerate(innings_blocks):
    title_tag = block.find("span")
    if not title_tag:
        continue

    innings_name = title_tag.get_text(strip=True).replace(" Innings", "")
    sheet_name = innings_name[:30].replace(" ", "_")
    print(f"\n📘 Processing: {innings_name}")

    ws = wb.create_sheet(title=sheet_name)

    rows = block.find_all("div", class_="cb-col cb-col-100 cb-scrd-itms")

    # --- Batting Data ---
    batting_data = []
    for row in rows:
        cols = row.find_all("div", recursive=False)
        cols_text = [c.get_text(strip=True) for c in cols if c.text.strip()]
        if len(cols_text) == 7 and cols_text[0].lower() not in ["extras", "total"]:
            batting_data.append(cols_text)

    if batting_data:
        df_bat = pd.DataFrame(batting_data, columns=["Batter", "Dismissal", "R", "B", "4s", "6s", "SR"])
        ws.append(["Batting Scorecard"])
        for r in dataframe_to_rows(df_bat, index=False, header=True):
            ws.append(r)

    # --- Extras and Total ---
    extras = ''
    total = ''
    for row in rows:
        cols = row.find_all("div", recursive=False)
        cols_text = [c.get_text(strip=True) for c in cols if c.text.strip()]
        if cols_text:
            if cols_text[0].lower() == "extras":
                extras = ' '.join(cols_text)
            elif cols_text[0].lower() == "total":
                total = ' '.join(cols_text)

    if extras:
        ws.append([])
        ws.append([extras])
    if total:
        ws.append([total])

    # --- Fall of Wickets ---
    fow_tag = block.find("div", class_="cb-col cb-col-100 cb-scrd-fall")
    if fow_tag:
        fow_text = fow_tag.get_text(strip=True).replace("Fall of wickets:", "")
        ws.append([])
        ws.append(["Fall of Wickets", fow_text])

    # --- Bowling Data (next block assumed) ---
    if i + 1 < len(innings_blocks):
        next_block = innings_blocks[i + 1]
        bowl_rows = next_block.find_all("div", class_="cb-col cb-col-100 cb-scrd-itms")
        bowling_data = []
        for row in bowl_rows:
            cols = row.find_all("div", recursive=False)
            cols_text = [c.get_text(strip=True) for c in cols if c.text.strip()]
            if len(cols_text) == 8:
                bowling_data.append(cols_text)

        if bowling_data:
            df_bowl = pd.DataFrame(bowling_data, columns=["Bowler", "O", "M", "R", "W", "NB", "WD", "ECO"])
            ws.append([])
            ws.append(["Bowling Scorecard"])
            for r in dataframe_to_rows(df_bowl, index=False, header=True):
                ws.append(r)

    # --- Formatting ---
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1 or cell.value in ["Batting Scorecard", "Bowling Scorecard", "Fall of Wickets"]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")

# Save the Excel workbook
excel_file = "Formatted_Match_Scorecard_records.xlsx"
wb.save(excel_file)
print(f"\n✅ All data saved in: {excel_file}")







